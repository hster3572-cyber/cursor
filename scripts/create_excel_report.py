#!/usr/bin/env python3
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference


def create_excel_report(model_dir: str):
    """Создает детальный Excel отчет с финансовой моделью"""
    
    # Читаем данные
    summary_df = pd.read_csv(os.path.join(model_dir, "department_summary.csv"))
    monthly_total_df = pd.read_csv(os.path.join(model_dir, "monthly_total.csv"))
    
    # Создаем Excel файл
    wb = Workbook()
    
    # Удаляем дефолтный лист
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Стили
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    percent_format = '0.0%'
    
    # === Лист 1: Резюме ===
    ws_summary = wb.create_sheet("Резюме")
    
    ws_summary['A1'] = "ФИНАНСОВАЯ МОДЕЛЬ ВНЕДРЕНИЯ AI-АССИСТЕНТА"
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A3'] = "Период анализа:"
    ws_summary['B3'] = "12 месяцев"
    ws_summary['A4'] = "Валюта:"
    ws_summary['B4'] = "USD"
    ws_summary['A5'] = "Дата анализа:"
    ws_summary['B5'] = "2025-10-20"
    
    # Ключевые финансовые показатели
    total_benefit = monthly_total_df['monthly_benefit'].sum()
    total_cost = monthly_total_df['monthly_cost'].sum()
    net_benefit = total_benefit - total_cost
    roi = (net_benefit / total_cost) if total_cost > 0 else 0
    
    ws_summary['A7'] = "КЛЮЧЕВЫЕ ФИНАНСОВЫЕ ПОКАЗАТЕЛИ"
    ws_summary['A7'].font = Font(bold=True, size=12)
    ws_summary.merge_cells('A7:D7')
    
    metrics = [
        ["Показатель", "Значение", "", ""],
        ["Суммарная выгода", total_benefit, "", ""],
        ["Суммарные затраты", total_cost, "", ""],
        ["Чистая выгода", net_benefit, "", ""],
        ["ROI (окупаемость инвестиций)", roi, "", ""],
        ["Средний срок окупаемости", "3-5 месяцев", "", ""],
    ]
    
    for row_idx, row_data in enumerate(metrics, start=8):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 8:  # Заголовок
                cell.fill = header_fill
                cell.font = header_font
            if col_idx == 2 and row_idx > 8 and row_idx <= 11:  # Финансовые значения
                cell.number_format = currency_format
            if row_idx == 12 and col_idx == 2:  # ROI
                cell.number_format = percent_format
    
    # === Лист 2: Сводка по отделам ===
    ws_dept = wb.create_sheet("Отделы")
    
    ws_dept['A1'] = "ДЕТАЛЬНЫЕ ПОКАЗАТЕЛИ ПО ОТДЕЛАМ (12 МЕСЯЦЕВ)"
    ws_dept['A1'].font = title_font
    ws_dept.merge_cells('A1:H1')
    
    dept_headers = ['Отдел', 'Пользователи', 'Годовая стоимость FTE', 'Выгода за 12 мес', 
                    'Затраты за 12 мес', 'Чистая выгода', 'ROI', 'Срок окупаемости']
    
    for col_idx, header in enumerate(dept_headers, start=1):
        cell = ws_dept.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, (_, row) in enumerate(summary_df.iterrows(), start=4):
        ws_dept.cell(row=row_idx, column=1, value=row['department'])
        ws_dept.cell(row=row_idx, column=2, value=int(row['eligible_users']))
        
        cell = ws_dept.cell(row=row_idx, column=3, value=row['annual_flc_per_fte'])
        cell.number_format = currency_format
        
        cell = ws_dept.cell(row=row_idx, column=4, value=row['month_12_benefit'])
        cell.number_format = currency_format
        
        cell = ws_dept.cell(row=row_idx, column=5, value=row['month_12_cost'])
        cell.number_format = currency_format
        
        cell = ws_dept.cell(row=row_idx, column=6, value=row['month_12_net'])
        cell.number_format = currency_format
        
        cell = ws_dept.cell(row=row_idx, column=7, value=row['roi'])
        cell.number_format = percent_format
        
        ws_dept.cell(row=row_idx, column=8, value=f"{row['payback_month']} мес")
    
    # Итого
    last_row = len(summary_df) + 4
    ws_dept.cell(row=last_row, column=1, value="ИТОГО").font = Font(bold=True)
    
    cell = ws_dept.cell(row=last_row, column=2, value=summary_df['eligible_users'].sum())
    cell.font = Font(bold=True)
    
    cell = ws_dept.cell(row=last_row, column=4, value=summary_df['month_12_benefit'].sum())
    cell.number_format = currency_format
    cell.font = Font(bold=True)
    
    cell = ws_dept.cell(row=last_row, column=5, value=summary_df['month_12_cost'].sum())
    cell.number_format = currency_format
    cell.font = Font(bold=True)
    
    cell = ws_dept.cell(row=last_row, column=6, value=summary_df['month_12_net'].sum())
    cell.number_format = currency_format
    cell.font = Font(bold=True)
    
    # === Лист 3: Помесячная детализация ===
    ws_monthly = wb.create_sheet("Помесячные данные")
    
    ws_monthly['A1'] = "ПОМЕСЯЧНАЯ ДИНАМИКА ЗАТРАТ И ВЫГОД"
    ws_monthly['A1'].font = title_font
    ws_monthly.merge_cells('A1:J1')
    
    monthly_headers = ['Месяц', 'Лицензии', 'Обучение', 'Управление изменениями', 
                       'Фикс. затраты орг.', 'Экономия админ. времени', 'Рост продуктивности',
                       'Суммарная выгода', 'Суммарные затраты', 'Чистая выгода']
    
    for col_idx, header in enumerate(monthly_headers, start=1):
        cell = ws_monthly.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, (_, row) in enumerate(monthly_total_df.iterrows(), start=4):
        ws_monthly.cell(row=row_idx, column=1, value=int(row['month']))
        
        for col_idx, col_name in enumerate(['license_cost', 'training_cost', 'change_mgmt_cost', 
                                            'org_fixed_cost', 'admin_time_savings', 
                                            'productivity_uplift_value', 'monthly_benefit', 
                                            'monthly_cost', 'net_benefit'], start=2):
            cell = ws_monthly.cell(row=row_idx, column=col_idx, value=row[col_name])
            cell.number_format = currency_format
    
    # Итого
    last_row = len(monthly_total_df) + 4
    ws_monthly.cell(row=last_row, column=1, value="ИТОГО").font = Font(bold=True)
    
    for col_idx, col_name in enumerate(['license_cost', 'training_cost', 'change_mgmt_cost', 
                                        'org_fixed_cost', 'admin_time_savings', 
                                        'productivity_uplift_value', 'monthly_benefit', 
                                        'monthly_cost', 'net_benefit'], start=2):
        cell = ws_monthly.cell(row=last_row, column=col_idx, value=monthly_total_df[col_name].sum())
        cell.number_format = currency_format
        cell.font = Font(bold=True)
    
    # === Лист 4: Кумулятивные показатели ===
    ws_cumulative = wb.create_sheet("Кумулятивные показатели")
    
    ws_cumulative['A1'] = "КУМУЛЯТИВНАЯ ДИНАМИКА"
    ws_cumulative['A1'].font = title_font
    ws_cumulative.merge_cells('A1:D1')
    
    cumulative_data = monthly_total_df.copy()
    cumulative_data['cumulative_benefit'] = cumulative_data['monthly_benefit'].cumsum()
    cumulative_data['cumulative_cost'] = cumulative_data['monthly_cost'].cumsum()
    cumulative_data['cumulative_net'] = cumulative_data['net_benefit'].cumsum()
    
    cum_headers = ['Месяц', 'Кумулятивная выгода', 'Кумулятивные затраты', 'Кумулятивная чистая выгода']
    
    for col_idx, header in enumerate(cum_headers, start=1):
        cell = ws_cumulative.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, (_, row) in enumerate(cumulative_data.iterrows(), start=4):
        ws_cumulative.cell(row=row_idx, column=1, value=int(row['month']))
        
        cell = ws_cumulative.cell(row=row_idx, column=2, value=row['cumulative_benefit'])
        cell.number_format = currency_format
        
        cell = ws_cumulative.cell(row=row_idx, column=3, value=row['cumulative_cost'])
        cell.number_format = currency_format
        
        cell = ws_cumulative.cell(row=row_idx, column=4, value=row['cumulative_net'])
        cell.number_format = currency_format
    
    # === Лист 5: Допущения ===
    ws_assumptions = wb.create_sheet("Допущения")
    
    ws_assumptions['A1'] = "КЛЮЧЕВЫЕ ДОПУЩЕНИЯ МОДЕЛИ"
    ws_assumptions['A1'].font = title_font
    ws_assumptions.merge_cells('A1:C1')
    
    assumptions = [
        ["Параметр", "Значение", "Примечание"],
        ["Стоимость лицензии", "$30", "за пользователя в месяц"],
        ["Фикс. затраты на внедрение", "$50,000", "единовременно"],
        ["Управление изменениями", "$5,000", "на отдел"],
        ["Обучение", "4 часа", "на пользователя"],
        ["Ставка дисконтирования", "10%", "годовых"],
        ["Монетизация продуктивности", "50%", "в первый год"],
        ["", "", ""],
        ["ПО ОТДЕЛАМ", "", ""],
        ["HR: Админ. время", "35%", "покрытие автоматизации 40%"],
        ["Finance: Админ. время", "30%", "покрытие автоматизации 35%"],
        ["IT: Админ. время", "25%", "покрытие автоматизации 30%"],
        ["Operations: Админ. время", "20%", "покрытие автоматизации 25%"],
        ["Legal/Admin: Админ. время", "40%", "покрытие автоматизации 35%"],
    ]
    
    for row_idx, row_data in enumerate(assumptions, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_assumptions.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:  # Заголовок
                cell.fill = header_fill
                cell.font = header_font
            if row_idx == 11 or row_idx == 4:  # Подзаголовок
                cell.font = Font(bold=True)
    
    # Настройка ширины колонок для всех листов
    for ws in wb.worksheets:
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row_idx, col_idx)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем файл
    output_path = os.path.join(os.path.dirname(model_dir), "HR_AI_Financial_Model.xlsx")
    wb.save(output_path)
    print(f"✅ Excel отчет сохранен: {output_path}")
    return output_path


if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.dirname(__file__))
    model_dir = os.path.join(base_dir, "outputs", "model")
    create_excel_report(model_dir)
